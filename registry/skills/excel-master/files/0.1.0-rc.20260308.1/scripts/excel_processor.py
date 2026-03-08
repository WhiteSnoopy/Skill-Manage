#!/usr/bin/env python3
"""
Excel Master Processor - 全能Excel智能处理引擎
支持通过JSON指令集批量处理Excel文件
"""

import json
import sys
from pathlib import Path
from typing import Any
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


class ExcelProcessor:
    """Excel处理引擎"""
    
    def __init__(self, input_file: str):
        self.input_path = Path(input_file)
        self.output_dir = self.input_path.parent
        self.df = None
        self.wb = None
        self.ws = None
        self.log = []
        
    def load_as_dataframe(self, sheet_name: str = None) -> pd.DataFrame:
        """加载为DataFrame"""
        kwargs = {'sheet_name': sheet_name} if sheet_name else {}
        self.df = pd.read_excel(self.input_path, **kwargs)
        return self.df
    
    def load_as_workbook(self) -> Workbook:
        """加载为openpyxl Workbook"""
        self.wb = load_workbook(self.input_path)
        self.ws = self.wb.active
        return self.wb
    
    def save(self, suffix: str = 'processed', output_path: str = None) -> Path:
        """保存文件"""
        if output_path:
            out = Path(output_path)
        else:
            out = self.output_dir / f"{self.input_path.stem}_{suffix}.xlsx"
        
        if self.wb:
            self.wb.save(out)
        elif self.df is not None:
            self.df.to_excel(out, index=False)
        
        return out
    
    # ========== 数据操作 ==========
    
    def filter_rows(self, column: str, condition: str, value: Any) -> int:
        """筛选行"""
        if self.df is None:
            self.load_as_dataframe()
        
        original_len = len(self.df)
        ops = {
            '>': lambda x: x > value,
            '>=': lambda x: x >= value,
            '<': lambda x: x < value,
            '<=': lambda x: x <= value,
            '==': lambda x: x == value,
            '!=': lambda x: x != value,
            'contains': lambda x: x.astype(str).str.contains(str(value), na=False),
            'startswith': lambda x: x.astype(str).str.startswith(str(value)),
            'endswith': lambda x: x.astype(str).str.endswith(str(value)),
        }
        
        if condition in ops:
            self.df = self.df[ops[condition](self.df[column])]
        
        filtered = original_len - len(self.df)
        self.log.append(f"筛选 {column} {condition} {value} → 过滤了 {filtered} 行")
        return filtered
    
    def sort_data(self, columns: list, ascending: list = None) -> None:
        """排序"""
        if self.df is None:
            self.load_as_dataframe()
        
        if ascending is None:
            ascending = [True] * len(columns)
        
        self.df = self.df.sort_values(by=columns, ascending=ascending)
        self.log.append(f"排序: {columns}, 升序: {ascending}")
    
    def remove_duplicates(self, columns: list = None) -> int:
        """去重"""
        if self.df is None:
            self.load_as_dataframe()
        
        original_len = len(self.df)
        self.df = self.df.drop_duplicates(subset=columns)
        removed = original_len - len(self.df)
        self.log.append(f"去重 {columns or '全部列'} → 移除 {removed} 行")
        return removed
    
    def replace_values(self, column: str, old_value: Any, new_value: Any) -> int:
        """替换值"""
        if self.df is None:
            self.load_as_dataframe()
        
        count = (self.df[column] == old_value).sum()
        self.df[column] = self.df[column].replace(old_value, new_value)
        self.log.append(f"替换 {column}: '{old_value}' → '{new_value}' ({count}处)")
        return count
    
    def remove_empty_rows(self) -> int:
        """删除空行"""
        if self.df is None:
            self.load_as_dataframe()
        
        original_len = len(self.df)
        self.df = self.df.dropna(how='all')
        removed = original_len - len(self.df)
        self.log.append(f"删除空行 → 移除 {removed} 行")
        return removed
    
    def split_column(self, column: str, delimiter: str, new_columns: list) -> None:
        """拆分列"""
        if self.df is None:
            self.load_as_dataframe()
        
        split_df = self.df[column].str.split(delimiter, expand=True)
        for i, new_col in enumerate(new_columns):
            if i < split_df.shape[1]:
                self.df[new_col] = split_df[i]
        
        self.log.append(f"拆分列 {column} → {new_columns}")
    
    def merge_columns(self, columns: list, new_column: str, separator: str = '') -> None:
        """合并列"""
        if self.df is None:
            self.load_as_dataframe()
        
        self.df[new_column] = self.df[columns].astype(str).agg(separator.join, axis=1)
        self.log.append(f"合并列 {columns} → {new_column}")
    
    # ========== 计算操作 ==========
    
    def add_sum_row(self, columns: list = None) -> None:
        """添加求和行"""
        if self.df is None:
            self.load_as_dataframe()
        
        if columns is None:
            columns = self.df.select_dtypes(include=['number']).columns.tolist()
        
        sum_row = {col: self.df[col].sum() if col in columns else '合计' for col in self.df.columns}
        self.df = pd.concat([self.df, pd.DataFrame([sum_row])], ignore_index=True)
        self.log.append(f"添加求和行: {columns}")
    
    def add_calculated_column(self, new_column: str, formula: str) -> None:
        """添加计算列（使用eval）"""
        if self.df is None:
            self.load_as_dataframe()
        
        # 替换列名为df['列名']格式
        for col in self.df.columns:
            formula = formula.replace(col, f"self.df['{col}']")
        
        self.df[new_column] = eval(formula)
        self.log.append(f"添加计算列 {new_column}")
    
    def add_excel_formula_column(self, column_letter: str, formula_template: str, 
                                  start_row: int = 2, end_row: int = None) -> None:
        """添加Excel公式列"""
        if self.wb is None:
            self.load_as_workbook()
        
        if end_row is None:
            end_row = self.ws.max_row
        
        for row in range(start_row, end_row + 1):
            formula = formula_template.replace('{row}', str(row))
            self.ws[f'{column_letter}{row}'] = formula
        
        self.log.append(f"添加公式列 {column_letter}: {formula_template}")
    
    # ========== 格式操作 ==========
    
    def format_header(self, bold: bool = True, bg_color: str = '4472C4', 
                      font_color: str = 'FFFFFF', center: bool = True) -> None:
        """格式化表头"""
        if self.wb is None:
            self.load_as_workbook()
        
        for cell in self.ws[1]:
            if bold:
                cell.font = Font(bold=True, color=font_color)
            if bg_color:
                cell.fill = PatternFill('solid', fgColor=bg_color)
            if center:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        self.log.append("格式化表头")
    
    def set_column_width(self, column_widths: dict) -> None:
        """设置列宽"""
        if self.wb is None:
            self.load_as_workbook()
        
        for col, width in column_widths.items():
            self.ws.column_dimensions[col].width = width
        
        self.log.append(f"设置列宽: {column_widths}")
    
    def auto_column_width(self) -> None:
        """自动调整列宽"""
        if self.wb is None:
            self.load_as_workbook()
        
        for column_cells in self.ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            self.ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        self.log.append("自动调整列宽")
    
    def add_conditional_format(self, range_str: str, rule_type: str, 
                                value: Any, fill_color: str = 'FFCCCC') -> None:
        """添加条件格式"""
        if self.wb is None:
            self.load_as_workbook()
        
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        if rule_type == 'lessThan':
            rule = CellIsRule(operator='lessThan', formula=[str(value)], fill=fill)
        elif rule_type == 'greaterThan':
            rule = CellIsRule(operator='greaterThan', formula=[str(value)], fill=fill)
        elif rule_type == 'equal':
            rule = CellIsRule(operator='equal', formula=[f'"{value}"' if isinstance(value, str) else str(value)], fill=fill)
        
        self.ws.conditional_formatting.add(range_str, rule)
        self.log.append(f"条件格式 {range_str}: {rule_type} {value}")
    
    def format_as_currency(self, columns: list, symbol: str = '¥') -> None:
        """格式化为货币"""
        if self.wb is None:
            self.load_as_workbook()
        
        for col in columns:
            for row in range(2, self.ws.max_row + 1):
                cell = self.ws[f'{col}{row}']
                cell.number_format = f'{symbol}#,##0.00'
        
        self.log.append(f"货币格式: {columns}")
    
    def format_as_percentage(self, columns: list, decimals: int = 1) -> None:
        """格式化为百分比"""
        if self.wb is None:
            self.load_as_workbook()
        
        fmt = f'0.{"0"*decimals}%'
        for col in columns:
            for row in range(2, self.ws.max_row + 1):
                cell = self.ws[f'{col}{row}']
                cell.number_format = fmt
        
        self.log.append(f"百分比格式: {columns}")
    
    # ========== 结构操作 ==========
    
    def add_row(self, position: int, values: list) -> None:
        """插入行"""
        if self.wb is None:
            self.load_as_workbook()
        
        self.ws.insert_rows(position)
        for col, value in enumerate(values, 1):
            self.ws.cell(row=position, column=col, value=value)
        
        self.log.append(f"插入行 {position}")
    
    def delete_rows(self, condition_column: str = None, condition_value: Any = None, 
                    empty_only: bool = False) -> int:
        """删除行"""
        if self.df is None:
            self.load_as_dataframe()
        
        original_len = len(self.df)
        
        if empty_only:
            self.df = self.df.dropna(how='all')
        elif condition_column and condition_value is not None:
            self.df = self.df[self.df[condition_column] != condition_value]
        
        removed = original_len - len(self.df)
        self.log.append(f"删除行 → 移除 {removed} 行")
        return removed
    
    def rename_columns(self, mapping: dict) -> None:
        """重命名列"""
        if self.df is None:
            self.load_as_dataframe()
        
        self.df = self.df.rename(columns=mapping)
        self.log.append(f"重命名列: {mapping}")
    
    def reorder_columns(self, new_order: list) -> None:
        """重排列顺序"""
        if self.df is None:
            self.load_as_dataframe()
        
        self.df = self.df[new_order]
        self.log.append(f"重排列顺序: {new_order}")
    
    # ========== 多表操作 ==========
    
    def merge_sheets(self, sheet_names: list, on: str, how: str = 'left') -> None:
        """合并多个sheet"""
        dfs = [pd.read_excel(self.input_path, sheet_name=name) for name in sheet_names]
        self.df = dfs[0]
        for df in dfs[1:]:
            self.df = pd.merge(self.df, df, on=on, how=how)
        
        self.log.append(f"合并sheets {sheet_names} on {on}")
    
    def split_by_column(self, column: str) -> dict:
        """按列值拆分为多个sheet"""
        if self.df is None:
            self.load_as_dataframe()
        
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        groups = self.df.groupby(column)
        sheet_info = {}
        
        for name, group in groups:
            ws = self.wb.create_sheet(title=str(name)[:31])
            for r_idx, row in enumerate(dataframe_to_rows(group, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            sheet_info[str(name)] = len(group)
        
        self.log.append(f"按 {column} 拆分为 {len(sheet_info)} 个sheet")
        return sheet_info
    
    # ========== 分析操作 ==========
    
    def create_pivot_table(self, values: str, index: list, columns: list = None,
                           aggfunc: str = 'sum') -> pd.DataFrame:
        """创建数据透视表"""
        if self.df is None:
            self.load_as_dataframe()
        
        pivot = pd.pivot_table(
            self.df,
            values=values,
            index=index,
            columns=columns,
            aggfunc=aggfunc,
            fill_value=0,
            margins=True,
            margins_name='合计'
        )
        
        self.df = pivot.reset_index()
        self.log.append(f"创建透视表: values={values}, index={index}, columns={columns}")
        return pivot
    
    def add_chart(self, chart_type: str, data_range: str, title: str = '',
                  position: str = 'E2') -> None:
        """添加图表"""
        if self.wb is None:
            self.load_as_workbook()
        
        chart_classes = {
            'bar': BarChart,
            'line': LineChart,
            'pie': PieChart
        }
        
        chart = chart_classes.get(chart_type, BarChart)()
        chart.title = title
        
        # 解析数据范围
        parts = data_range.split(':')
        min_col = ord(parts[0][0].upper()) - ord('A') + 1
        min_row = int(parts[0][1:])
        max_col = ord(parts[1][0].upper()) - ord('A') + 1
        max_row = int(parts[1][1:])
        
        data = Reference(self.ws, min_col=min_col, min_row=min_row, 
                        max_col=max_col, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        
        self.ws.add_chart(chart, position)
        self.log.append(f"添加{chart_type}图表: {data_range}")
    
    def get_summary(self) -> dict:
        """获取数据摘要"""
        if self.df is None:
            self.load_as_dataframe()
        
        return {
            'rows': len(self.df),
            'columns': len(self.df.columns),
            'column_names': self.df.columns.tolist(),
            'dtypes': self.df.dtypes.astype(str).to_dict(),
            'null_counts': self.df.isnull().sum().to_dict()
        }
    
    def get_processing_log(self) -> list:
        """获取处理日志"""
        return self.log


def process_from_json(input_file: str, operations_json: str, output_suffix: str = 'processed'):
    """从JSON指令处理Excel"""
    processor = ExcelProcessor(input_file)
    operations = json.loads(operations_json)
    
    for op in operations:
        method = op.get('method')
        params = op.get('params', {})
        
        if hasattr(processor, method):
            getattr(processor, method)(**params)
    
    output_path = processor.save(suffix=output_suffix)
    
    return {
        'output_file': str(output_path),
        'log': processor.get_processing_log(),
        'summary': processor.get_summary()
    }


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python excel_processor.py <input_file> --operations '<json_operations>'")
        print("       python excel_processor.py <input_file> --info")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    if sys.argv[2] == '--info':
        processor = ExcelProcessor(input_file)
        summary = processor.get_summary()
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    elif sys.argv[2] == '--operations':
        operations_json = sys.argv[3]
        result = process_from_json(input_file, operations_json)
        print(json.dumps(result, ensure_ascii=False, indent=2))
